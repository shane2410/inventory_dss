from django.shortcuts import render, redirect, get_object_or_404
from functools import wraps
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum
import math
import re
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from .models import Product, Material, SalesData, Transaction, BOM, ProductRatio, DisaggregatedPlan, CustomerOrder, MPSConfiguration, SelectedProductForMPS
from .forms import ImportDataForm, MonthlyForecastImportForm, TransactionForm
from .services import aggregate_material_demand, abc_classification, disaggregate_forecast, forecast_monthly_total, forecast_product, forecast_product_monthly, run_dss, get_demand_by_product, get_orders_by_product, ppa_lot_sizing, calculate_mps
from .recommendations import build_dashboard_recommendations, build_inventory_alert_recommendations, build_inventory_watchlist_recommendations
from datetime import datetime, timedelta, date
from .permissions import (
    ALL_ROLE_CODES,
    ROLE_ADMIN,
    ROLE_FORM_CHOICES,
    ROLE_MANAGER,
    ROLE_OTHER,
    ROLE_STAFF,
    assign_user_role,
    create_user_with_role,
    ensure_role_groups,
    get_user_role_code,
    get_user_role_label,
    role_required,
)


def allocate_overtime(shortage, regular, alpha_percent=20):
    """
    Phân bổ tăng ca (OT) tập trung vào giữa shortage window

    Parameters:
    - shortage: list[int/float]  (thiếu hàng theo tháng)
    - regular:  list[int/float]  (sản lượng giờ thường)
    - alpha_percent: int/float   (% OT max, ví dụ 20)

    Returns:
    - OT: list[float] (tăng ca từng tháng)
    """

    # ===== 1. Validate input =====
    if len(shortage) != len(regular):
        raise ValueError("shortage và regular phải cùng độ dài")

    if alpha_percent < 0 or alpha_percent > 50:
        raise ValueError("alpha nên nằm trong khoảng 0–50 (%)")

    alpha = alpha_percent / 100
    n = len(shortage)
    OT = np.zeros(n)

    # ===== 2. Xác định shortage window =====
    shortage_indices = [i for i in range(n) if shortage[i] > 0]

    if not shortage_indices:
        return OT.tolist()
    start = min(shortage_indices)
    end = max(shortage_indices)

    # ===== 3. Tính center =====
    center = (start + end) / 2

    # ===== 4. Tạo weight (Gaussian mượt hơn) =====
    weights = np.zeros(n)
    sigma = max((end - start) / 2, 1)  # tránh chia 0

    for t in range(start, end + 1):
        distance = abs(t - center)
        weights[t] = np.exp(-(distance ** 2) / (2 * sigma ** 2))

    # Normalize weights
    total_weight = weights.sum()
    if total_weight == 0:
        return OT.tolist()

    weights = weights / total_weight

    # ===== 5. Tổng shortage cần bù =====
    total_shortage = sum(shortage)

    # ===== 6. Phân bổ OT =====
    for t in range(start, end + 1):
        max_ot = alpha * regular[t]
        ot_alloc = weights[t] * total_shortage

        OT[t] = min(ot_alloc, max_ot)

    return OT.tolist()


def is_feasible(demand, regular, alpha):
    """
    Kiểm tra khả năng thực hiện kế hoạch.

    In cảnh báo nếu tổng capacity không đủ đáp ứng tổng demand.
    
    Parameters:
    - demand: list (tổng nhu cầu)
    - regular: list (sản lượng giờ thường theo tháng)
    - alpha: float (% OT max so với regular)
    
    Returns:
    - bool: True nếu khả thi, False nếu không
    """
    total_capacity = sum(regular) + sum(alpha * r for r in regular)
    total_demand = sum(demand)

    if total_capacity < total_demand:
        print("❌ KHÔNG KHẢ THI")
        return False

    return True


def optimize_aggregate_plan_lp(demand, regular_caps, cost_params, alpha=0.2, safety_stock=0):
    """
    Giải Aggregate Planning bằng Linear Programming.
    
    Parameters:
    - demand: list (nhu cầu từng tháng)
    - regular_caps: list (capacity giờ thường)
    - cost_params: dict (regular_cost, overtime_cost, subcontract_cost, inventory_cost, backorder_cost)
    - alpha: float (OT limit, e.g., 0.2 = 20%)
    - safety_stock: float (minimum inventory level, not for final period)
    
    Returns:
    - result: list[dict] với keys: regular, overtime, subcontract, inventory, backlog cho mỗi tháng
    - status hoặc total_cost
    """
    import pulp
    
    n = len(demand)
    model = pulp.LpProblem("Aggregate_Planning", pulp.LpMinimize)
    
    # ===== Variables =====
    R = pulp.LpVariable.dicts("R", range(n), lowBound=0)
    O = pulp.LpVariable.dicts("O", range(n), lowBound=0)
    S = pulp.LpVariable.dicts("S", range(n), lowBound=0)
    I = pulp.LpVariable.dicts("I", range(n), lowBound=0)
    B = pulp.LpVariable.dicts("B", range(n), lowBound=0)
    
    # ===== Objective: Minimize total cost =====
    model += pulp.lpSum(
        R[t] * cost_params['regular_cost']
        + O[t] * cost_params['overtime_cost']
        + S[t] * cost_params['subcontract_cost']
        + I[t] * cost_params['inventory_cost']
        + B[t] * cost_params['backorder_cost']
        for t in range(n)
    )
    
    # ===== Constraints =====
    for t in range(n):
        # Inventory balance equation
        if t == 0:
            model += R[t] + O[t] + S[t] == demand[t] + I[t] - B[t]
        else:
            model += I[t-1] + R[t] + O[t] + S[t] == demand[t] + B[t-1] + I[t]
        
        # Capacity constraints
        model += R[t] <= regular_caps[t]
        
        # Overtime capacity
        if cost_params['overtime_cost'] > 0:
            model += O[t] <= alpha * regular_caps[t]
        else:
            model += O[t] == 0
        
        # Subcontract capacity
        if cost_params['subcontract_cost'] <= 0:
            model += S[t] == 0
        
        # Safety stock constraint (not for final period)
        if safety_stock > 0 and t < n - 1:
            model += I[t] >= safety_stock
    
    # Final constraint: Ending inventory = 0, No backlog in final period
    model += I[n-1] == 0
    model += B[n-1] == 0
    
    # ===== Solve =====
    model.solve(pulp.PULP_CBC_CMD(msg=0))
    
    # Check if solution exists
    status = pulp.LpStatus[model.status]
    if status != 'Optimal':
        return None, status
    
    # ===== Extract results =====
    result = []
    for t in range(n):
        result.append({
            'regular': R[t].value() or 0,
            'overtime': O[t].value() or 0,
            'subcontract': S[t].value() or 0,
            'inventory': I[t].value() or 0,
            'backlog': B[t].value() or 0,
        })
    
    total_cost = pulp.value(model.objective)
    
    return result, total_cost


def find_best_ot_alpha(demand_rows, regular_caps, cost_params, safety_stock=0):
    """
    Tìm OT limit tối ưu (0-50%) với min total cost.
    Chạy LP mỗi lần thử alpha để tìm min cost solution.
    
    Parameters:
    - demand_rows: list (nhu cầu mỗi tháng)
    - regular_caps: list (capacity giờ thường mỗi tháng)
    - cost_params: dict (regular_cost, overtime_cost, subcontract_cost, inventory_cost, backorder_cost)
    - safety_stock: float (minimum inventory level)
    
    Returns:
    - best_alpha: float (% OT tối ưu, từ 0 đến 0.5)
    - best_cost: float (tổng cost tối ưu)
    """
    best_cost = float('inf')
    best_alpha = 0.0
    
    # Thử từ 0% → 50% (bước 2%)
    for alpha_pct in range(0, 51, 2):
        alpha = alpha_pct / 100.0
        
        # Chạy LP với alpha này
        result, total_cost = optimize_aggregate_plan_lp(demand_rows, regular_caps, cost_params, alpha=alpha, safety_stock=safety_stock)
        
        # Nếu không có feasible solution, skip
        if result is None:
            continue
        
        if total_cost < best_cost:
            best_cost = total_cost
            best_alpha = alpha
    
    return best_alpha, best_cost








@role_required(ROLE_ADMIN, ROLE_MANAGER, ROLE_STAFF)
def plan_synthesis(request):
    from .models import MonthlyProductionData

    def _to_float(raw_value, default=0.0):
        try:
            if raw_value is None:
                return float(default)
            text = str(raw_value).strip().replace(',', '')
            if text == '':
                return float(default)
            return float(text)
        except (TypeError, ValueError):
            return float(default)

    def _to_int(raw_value, default=0):
        try:
            if raw_value is None:
                return int(default)
            text = str(raw_value).strip().replace(',', '')
            if text == '':
                return int(default)
            return int(float(text))
        except (TypeError, ValueError):
            return int(default)

    def _add_months(base_month, months):
        month_index = base_month.month - 1 + months
        year = base_month.year + month_index // 12
        month = month_index % 12 + 1
        return base_month.replace(year=year, month=month, day=1)

    history_qs = MonthlyProductionData.objects.filter(
        source=MonthlyProductionData.SOURCE_PLANNING
    ).order_by('month')

    forecast_mean, forecast_std, forecast_8, mae, rmse, mape = forecast_monthly_total(history_qs=history_qs)

    forecast_rows = []
    if history_qs.exists():
        last_month = history_qs.last().month
        for idx, value in enumerate(forecast_8, start=1):
            forecast_rows.append({
                'month': _add_months(last_month, idx).strftime('%m/%Y'),
                'quantity': int(math.ceil(float(value or 0))),
            })
    else:
        for idx, value in enumerate(forecast_8, start=1):
            forecast_rows.append({
                'month': f'Tháng {idx}',
                'quantity': int(math.ceil(float(value or 0))),
            })

    defaults = {
        'opening_inventory': 0,
        'workers': 50,
        'productivity': 150,
        'regular_cost': 0,
        'overtime_cost': 0,
        'subcontract_cost': 0,
        'inventory_cost': 0,
        'backorder_cost': 0,
        'ot_limit_pct': 20,
        'inventory_policy': 0,
        'hire_cost': 0,
        'layoff_cost': 0,
        'safety_stock': 0,
    }

    workforce_adjustments = [0 for _ in forecast_rows]

    if request.method == 'POST':
        for key in defaults:
            if key == 'workers':
                defaults[key] = _to_int(request.POST.get(key), defaults[key])
            else:
                defaults[key] = _to_float(request.POST.get(key), defaults[key])

        for idx in range(len(forecast_rows)):
            workforce_adjustments[idx] = _to_int(request.POST.get(f'workforce_adjustment_{idx + 1}'), 0)

    opening_inventory = float(defaults['opening_inventory'])
    workers = int(defaults['workers'])
    productivity = float(defaults['productivity'])
    regular_cost = float(defaults['regular_cost'])
    overtime_cost = float(defaults['overtime_cost'])
    subcontract_cost = float(defaults['subcontract_cost'])
    inventory_cost = float(defaults['inventory_cost'])
    backorder_cost = float(defaults['backorder_cost'])
    ot_limit_pct = max(0.0, float(defaults['ot_limit_pct']))
    inventory_policy = max(0.0, float(defaults['inventory_policy']))
    hire_cost = float(defaults['hire_cost'])
    layoff_cost = float(defaults['layoff_cost'])
    safety_stock = max(0.0, float(defaults['safety_stock']))
    
    # Debug: log parameter values
    import sys
    print(f"DEBUG: overtime_cost={overtime_cost}, type={type(overtime_cost)}", file=sys.stderr)

    plan_rows = []
    running_inventory = int(math.ceil(opening_inventory))
    running_workers = float(workers)

    demand_rows = [int(row['quantity']) for row in forecast_rows]
    regular_caps = []
    overtime_caps = []
    worker_levels = []

    worker_tracker = float(workers)
    for idx in range(len(demand_rows)):
        worker_tracker = max(0.0, worker_tracker + workforce_adjustments[idx])
        worker_levels.append(worker_tracker)
        month_regular_capacity = max(0, int(math.floor(worker_tracker * productivity + 1e-9)))
        month_overtime_capacity = 0
        if overtime_cost > 0:
            month_overtime_capacity = max(0, int(math.floor(month_regular_capacity * ot_limit_pct / 100.0 + 1e-9)))
        regular_caps.append(month_regular_capacity)
        overtime_caps.append(month_overtime_capacity)

    # === Setup cost optimization ===
    alpha_ot = ot_limit_pct / 100.0  # convert % to decimal
    
    cost_params = {
        'regular_cost': regular_cost,
        'overtime_cost': overtime_cost,
        'subcontract_cost': subcontract_cost,
        'inventory_cost': inventory_cost,
        'backorder_cost': backorder_cost,
    }
    
    # Initialize feasibility tracking
    feasibility_ok = False
    feasibility_message = None
    
    # Nếu ot_limit_pct = 0 → tìm optimal alpha
    if ot_limit_pct <= 0:
        alpha_ot, _ = find_best_ot_alpha(demand_rows, regular_caps, cost_params, safety_stock=safety_stock)
        ot_limit_pct = alpha_ot * 100.0
        # Recalculate OT capacity với alpha tối ưu
        for idx in range(len(demand_rows)):
            month_regular_capacity = regular_caps[idx]
            month_overtime_capacity = max(0, int(math.floor(month_regular_capacity * ot_limit_pct / 100.0 + 1e-9)))
            overtime_caps[idx] = month_overtime_capacity
    
    # === Run LP optimization ===
    lp_result, status_or_cost = optimize_aggregate_plan_lp(demand_rows, regular_caps, cost_params, alpha=alpha_ot, safety_stock=safety_stock)
    
    # Initialize totals (before checking feasibility)
    total_regular = 0.0
    total_overtime = 0.0
    total_subcontract = 0.0
    total_inventory_cost = 0.0
    total_backorder_cost = 0.0
    total_regular_cost = 0.0
    total_overtime_cost = 0.0
    total_subcontract_cost = 0.0
    total_workforce_hire_cost = 0.0
    total_workforce_layoff_cost = 0.0
    total_cost = 0.0
    
    # Kiểm tra nếu LP không có feasible solution
    if lp_result is None:
        # Return error page with status message
        return render(request, 'inventory/plan_synthesis.html', {
            'forecast_rows': forecast_rows,
            'defaults': defaults,
            'plan_rows': [],
            'summary_cards': [],
            'history_rows': [],
            'planning_metrics': [],
            'cost_breakdown': [],
            'workforce_summary': [],
            'error_message': f'Kế hoạch không khả thi ({status_or_cost}). Hãy tăng OT limit hoặc bật subcontract.',
            'feasibility_ok': False,
        })
    
    # LP solved successfully
    lp_cost = status_or_cost
    feasibility_ok = True
    
    # === Build plan_rows từ LP result ===
    plan_rows = []
    
    for idx, demand_qty in enumerate(demand_rows):
        workforce_change = workforce_adjustments[idx] if idx < len(workforce_adjustments) else 0
        running_workers = worker_levels[idx]
        month_regular_capacity = regular_caps[idx]
        month_overtime_capacity = overtime_caps[idx]
        
        # Get LP solution for this month
        lp_month = lp_result[idx]
        regular_qty = int(round(lp_month['regular']))
        overtime_qty = int(round(lp_month['overtime']))
        subcontract_qty = int(round(lp_month['subcontract']))
        ending_inventory = int(round(lp_month['inventory']))
        backorder_qty = int(round(lp_month['backlog']))
        
        total_production = regular_qty + overtime_qty + subcontract_qty
        
        # Tính beginning inventory từ LP
        if idx == 0:
            beginning_inventory = int(math.ceil(opening_inventory))
        else:
            beginning_inventory = int(round(lp_result[idx-1]['inventory']))
        
        average_inventory = max(0.0, (beginning_inventory + ending_inventory) / 2.0)
        policy_gap = max(0.0, ending_inventory - inventory_policy)
        workforce_hire_cost = max(0, workforce_change) * hire_cost
        workforce_layoff_cost = max(0, -workforce_change) * layoff_cost
        
        row_regular_cost = regular_qty * regular_cost
        row_overtime_cost = overtime_qty * overtime_cost
        row_subcontract_cost = subcontract_qty * subcontract_cost
        row_inventory_cost = average_inventory * inventory_cost
        row_backorder_cost = backorder_qty * backorder_cost
        row_total_cost = (
            row_regular_cost
            + row_overtime_cost
            + row_subcontract_cost
            + row_inventory_cost
            + row_backorder_cost
            + workforce_hire_cost
            + workforce_layoff_cost
        )
        
        plan_rows.append({
            'month': str(forecast_rows[idx]['month']),
            'forecast': demand_qty,
            'beginning_inventory': beginning_inventory,
            'workers': running_workers,
            'workforce_change': workforce_change,
            'regular_capacity': month_regular_capacity,
            'overtime_capacity': month_overtime_capacity,
            'regular': regular_qty,
            'overtime': overtime_qty,
            'subcontract': subcontract_qty,
            'production': total_production,
            'ending_inventory': ending_inventory,
            'backorder': backorder_qty,
            'average_inventory': average_inventory,
            'policy_gap': policy_gap,
            'within_ot_limit': overtime_qty <= month_overtime_capacity + 1e-9,
            'regular_cost': row_regular_cost,
            'overtime_cost': row_overtime_cost,
            'subcontract_cost': row_subcontract_cost,
            'inventory_cost': row_inventory_cost,
            'backorder_cost': row_backorder_cost,
            'workforce_hire_cost': workforce_hire_cost,
            'workforce_layoff_cost': workforce_layoff_cost,
            'total_cost': row_total_cost,
        })
    
    # === Calculate totals ===
    total_regular = sum(row['regular'] for row in plan_rows)
    total_overtime = sum(row['overtime'] for row in plan_rows)
    total_subcontract = sum(row['subcontract'] for row in plan_rows)
    total_inventory_cost = sum(row['inventory_cost'] for row in plan_rows)
    total_backorder_cost = sum(row['backorder_cost'] for row in plan_rows)
    total_regular_cost = sum(row['regular_cost'] for row in plan_rows)
    total_overtime_cost = sum(row['overtime_cost'] for row in plan_rows)
    total_subcontract_cost = sum(row['subcontract_cost'] for row in plan_rows)
    total_workforce_hire_cost = sum(row['workforce_hire_cost'] for row in plan_rows)
    total_workforce_layoff_cost = sum(row['workforce_layoff_cost'] for row in plan_rows)
    total_cost = sum(row['total_cost'] for row in plan_rows)

    # === Build summary cards ===
    workforce_total_change = sum(workforce_adjustments)
    workforce_end = workers + workforce_total_change
    workforce_cost = total_workforce_hire_cost + total_workforce_layoff_cost
    workforce_action = 'Ổn định'
    if workforce_total_change > 0:
        workforce_action = 'Tuyển thêm'
    elif workforce_total_change < 0:
        workforce_action = 'Sa thải'

    summary_cards = [
        {
            'label': 'Tổng chi phí',
            'value': int(round(total_cost)),
        },
        {
            'label': 'Sản lượng thường',
            'value': int(total_regular),
        },
        {
            'label': 'Làm thêm giờ',
            'value': int(total_overtime),
        },
        {
            'label': 'Thuê ngoài',
            'value': int(total_subcontract),
        },
    ]

    history_rows = [
        {
            'month': item.month.strftime('%m/%Y'),
            'quantity': float(item.quantity or 0),
        }
        for item in history_qs
    ]

    planning_metrics = [
        {
            'label': 'DỰ BÁO',
            'is_group': False,
            'integer_values': True,
            'values': [int(row['forecast']) for row in plan_rows],
            'total': int(sum(row['forecast'] for row in plan_rows)),
        },
        {
            'label': 'SẢN LƯỢNG',
            'is_group': True,
            'items': [
                {
                    'label': 'Giờ thường',
                    'integer_values': True,
                    'values': [int(row['regular']) for row in plan_rows],
                    'total': int(total_regular),
                },
                {
                    'label': 'Tăng ca',
                    'integer_values': True,
                    'values': [int(row['overtime']) for row in plan_rows],
                    'total': int(total_overtime),
                },
                {
                    'label': 'Thuê ngoài',
                    'integer_values': True,
                    'values': [int(row['subcontract']) for row in plan_rows],
                    'total': int(total_subcontract),
                },
            ]
        },
        {
            'label': 'SẢN LƯỢNG – DỰ BÁO',
            'is_group': False,
            'integer_values': True,
            'values': [int(row['production'] - row['forecast']) for row in plan_rows],
            'total': int(sum((row['production'] - row['forecast']) for row in plan_rows)),
        },
        {
            'label': 'TỒN KHO',
            'is_group': True,
            'items': [
                {
                    'label': 'Tồn đầu kỳ',
                    'integer_values': True,
                    'values': [int(row['beginning_inventory']) for row in plan_rows],
                    'total': None,
                },
                {
                    'label': 'Tồn cuối kỳ',
                    'integer_values': True,
                    'values': [int(row['ending_inventory']) for row in plan_rows],
                    'total': None,
                },
                {
                    'label': 'Tồn trung bình',
                    'integer_values': True,
                    'values': [int(row['average_inventory']) for row in plan_rows],
                    'total': None,
                },
                {
                    'label': 'Thiếu hàng',
                    'integer_values': True,
                    'values': [int(row['backorder']) for row in plan_rows],
                    'total': int(sum(row['backorder'] for row in plan_rows)),
                },
            ]
        },
        {
            'label': 'CHI PHÍ',
            'is_group': True,
            'items': [
                {
                    'label': 'Chi phí sản xuất',
                    'is_subgroup': True,
                    'items': [
                        {
                            'label': 'Giờ thường',
                            'integer_values': True,
                            'values': [int(round(row['regular_cost'])) for row in plan_rows],
                            'total': int(round(total_regular_cost)),
                        },
                        {
                            'label': 'Tăng ca',
                            'integer_values': True,
                            'values': [int(round(row['overtime_cost'])) for row in plan_rows],
                            'total': int(round(total_overtime_cost)),
                        },
                        {
                            'label': 'Thuê ngoài',
                            'integer_values': True,
                            'values': [int(round(row['subcontract_cost'])) for row in plan_rows],
                            'total': int(round(total_subcontract_cost)),
                        },
                        {
                            'label': 'Tuyển dụng/Sa thải',
                            'editable': True,
                            'values': [int(row['workforce_change']) for row in plan_rows],
                            'total': int(round(sum(row['workforce_change'] for row in plan_rows))),
                        },
                    ]
                },
                {
                    'label': 'Chi phí tồn kho',
                    'integer_values': True,
                    'values': [int(round(row['inventory_cost'])) for row in plan_rows],
                    'total': int(round(total_inventory_cost)),
                },
                {
                    'label': 'Chi phí thiếu hàng',
                    'integer_values': True,
                    'values': [int(round(row['backorder_cost'])) for row in plan_rows],
                    'total': int(round(total_backorder_cost)),
                },
                {
                    'label': 'Tổng chi phí',
                    'integer_values': True,
                    'values': [int(round(row['total_cost'])) for row in plan_rows],
                    'total': int(round(total_cost)),
                },
            ]
        },
    ]

    first_plan_row = plan_rows[0] if plan_rows else {}

    return render(request, 'inventory/plan_synthesis.html', {
        'title': 'Kế hoạch tổng hợp',
        'history_rows': history_rows,
        'forecast_rows': forecast_rows,
        'plan_rows': plan_rows,
        'planning_metrics': planning_metrics,
        'summary_cards': summary_cards,
        'forecast_mean': round(float(forecast_mean or 0), 2),
        'forecast_std': round(float(forecast_std or 0), 2),
        'mae': round(float(mae or 0), 2),
        'rmse': round(float(rmse or 0), 2),
        'mape': round(float(mape or 0), 2),
        'opening_inventory': opening_inventory,
        'workers': workers,
        'productivity': productivity,
        'regular_capacity': first_plan_row.get('regular_capacity', 0.0),
        'overtime_capacity': first_plan_row.get('overtime_capacity', 0.0),
        'regular_cost': regular_cost,
        'overtime_cost': overtime_cost,
        'subcontract_cost': subcontract_cost,
        'inventory_cost': inventory_cost,
        'backorder_cost': backorder_cost,
        'ot_limit_pct': ot_limit_pct,
        'inventory_policy': inventory_policy,
        'hire_cost': hire_cost,
        'layoff_cost': layoff_cost,
        'safety_stock': safety_stock,
        'workforce_action': workforce_action,
        'workforce_cost': workforce_cost,
        'workforce_total_change': workforce_total_change,
        'workforce_end': workforce_end,
        'workforce_adjustments': workforce_adjustments,
        'feasibility_ok': feasibility_ok,
        'feasibility_message': None if feasibility_ok else '❌ KHÔNG KHẢ THI: tổng capacity hiện tại nhỏ hơn tổng demand, nên kế hoạch vẫn thiếu hàng.',
        'error_message': None,
        'forecast_note': 'OT limit = OT_t <= alpha x regular capacity. If overtime cost is blank, overtime is disabled. If subcontract cost is blank, subcontract is disabled. Inventory policy is treated as an upper limit I_t <= I_max. Production quantities are rounded to whole units.',
        'forecast_total': round(sum(item['quantity'] for item in forecast_rows), 2),
    })


@role_required(ROLE_ADMIN, ROLE_MANAGER, ROLE_STAFF)
def save_planning_config(request):
    """API endpoint to save planning configuration"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    from .models import PlanningConfiguration
    import json

    def _safe_float(value, default=0.0):
        try:
            if value is None:
                return float(default)
            text = str(value).strip().replace(',', '')
            if text == '':
                return float(default)
            return float(text)
        except (TypeError, ValueError):
            return float(default)

    def _safe_int(value, default=0):
        try:
            if value is None:
                return int(default)
            text = str(value).strip().replace(',', '')
            if text == '':
                return int(default)
            return int(float(text))
        except (TypeError, ValueError):
            return int(default)
    
    try:
        data = json.loads(request.body)
        
        # Check if config exists
        config = PlanningConfiguration.objects.last()
        is_update = config is not None
        
        # Save or update
        if config:
            config.opening_inventory = _safe_float(data.get('opening_inventory', 0))
            config.workers = _safe_int(data.get('workers', 50))
            config.productivity = _safe_float(data.get('productivity', 150))
            config.regular_cost = _safe_float(data.get('regular_cost', 0))
            config.overtime_cost = _safe_float(data.get('overtime_cost', 0))
            config.subcontract_cost = _safe_float(data.get('subcontract_cost', 0))
            config.inventory_cost = _safe_float(data.get('inventory_cost', 0))
            config.backorder_cost = _safe_float(data.get('backorder_cost', 0))
            config.ot_limit_pct = _safe_float(data.get('ot_limit_pct', 20))
            config.inventory_policy = _safe_float(data.get('inventory_policy', 0))
            config.current_workers = _safe_int(data.get('current_workers', 50))
            config.hire_cost = _safe_float(data.get('hire_cost', 0))
            config.layoff_cost = _safe_float(data.get('layoff_cost', 0))
            config.save()
            action = 'updated'
        else:
            config = PlanningConfiguration.objects.create(
                opening_inventory=_safe_float(data.get('opening_inventory', 0)),
                workers=_safe_int(data.get('workers', 50)),
                productivity=_safe_float(data.get('productivity', 150)),
                regular_cost=_safe_float(data.get('regular_cost', 0)),
                overtime_cost=_safe_float(data.get('overtime_cost', 0)),
                subcontract_cost=_safe_float(data.get('subcontract_cost', 0)),
                inventory_cost=_safe_float(data.get('inventory_cost', 0)),
                backorder_cost=_safe_float(data.get('backorder_cost', 0)),
                ot_limit_pct=_safe_float(data.get('ot_limit_pct', 20)),
                inventory_policy=_safe_float(data.get('inventory_policy', 0)),
                current_workers=_safe_int(data.get('current_workers', 50)),
                hire_cost=_safe_float(data.get('hire_cost', 0)),
                layoff_cost=_safe_float(data.get('layoff_cost', 0)),
            )
            action = 'created'
        
        return JsonResponse({
            'success': True,
            'action': action,
            'message': f'Dữ liệu kế hoạch đã được lưu thành công!'
        })
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@role_required(ROLE_ADMIN, ROLE_MANAGER, ROLE_STAFF)
def product_decomposition(request):
    start_month = '2025-05-01'
    forecast_mean, forecast_std, forecast_list, mae, rmse, mape = forecast_monthly_total()
    months = list(pd.date_range(start=start_month, periods=len(forecast_list), freq='MS'))

    forecast_map = {
        month.strftime('%Y_%m'): float(forecast_list[index] or 0)
        for index, month in enumerate(months)
    }

    # Filter by year+month instead of exact date to catch all records in each month
    from django.db.models import Q
    q_filter = Q()
    for month in months:
        q_filter |= Q(month__year=month.year, month__month=month.month)
    ratio_qs = ProductRatio.objects.filter(q_filter)
    
    # Build ratio_map with string keys (product_code_YYYY_MM) and ratio values
    ratio_map = {}
    for ratio in ratio_qs:
        key = f'{ratio.product_code}_{ratio.month.strftime("%Y_%m")}'
        ratio_map[key] = ratio.ratio

    existing_product_codes = []
    for ratio in ProductRatio.objects.values('product_code', 'product_name').order_by('product_code').distinct():
        existing_product_codes.append({
            'product_code': ratio['product_code'],
            'product_name': ratio['product_name'],
        })

    if not existing_product_codes:
        existing_product_codes = [
            {'product_code': '', 'product_name': ''},
            {'product_code': '', 'product_name': ''},
            {'product_code': '', 'product_name': ''},
        ]

    row_indices = list(range(len(existing_product_codes)))

    if request.method == 'POST':
        month_warnings = []

        posted_indices = set()
        for key in request.POST.keys():
            match = re.match(r'^product_code_(\d+)$', key)
            if match:
                posted_indices.add(int(match.group(1)))

        for row_index in sorted(posted_indices):
            original_code = (request.POST.get(f'original_code_{row_index}') or '').strip()
            product_code = (request.POST.get(f'product_code_{row_index}') or '').strip()
            product_name = (request.POST.get(f'product_name_{row_index}') or '').strip()

            row_has_values = bool(product_code or product_name)
            month_ratio_inputs = {}
            for month in months:
                month_key = month.strftime('%Y_%m')
                raw_ratio = (request.POST.get(f'ratio_{row_index}_{month_key}') or '').strip()
                month_ratio_inputs[month_key] = raw_ratio
                if raw_ratio:
                    row_has_values = True

            if original_code and not product_code:
                ProductRatio.objects.filter(product_code=original_code, month__in=months).delete()
                continue

            if not row_has_values:
                continue

            if not product_code:
                month_warnings.append(f'Hàng {row_index + 1}: cần nhập ID_P')
                continue

            if not product_name:
                product_name = product_code

            if original_code and original_code != product_code:
                ProductRatio.objects.filter(product_code=original_code, month__in=months).delete()

            month_ratio_total = 0.0
            for month in months:
                month_key = month.strftime('%Y_%m')
                total_forecast = forecast_map[month_key]
                raw_ratio = month_ratio_inputs[month_key]

                try:
                    ratio = float(raw_ratio) if raw_ratio else 0.0
                except (TypeError, ValueError):
                    ratio = 0.0

                if ratio < 0:
                    ratio = 0.0
                if ratio > 1:
                    ratio = 1.0

                forecast_qty = round(total_forecast * ratio, 2)
                ProductRatio.objects.update_or_create(
                    product_code=product_code,
                    month=month,
                    defaults={
                        'product_name': product_name,
                        'ratio': ratio,
                        'forecast_qty': forecast_qty,
                    }
                )
                month_ratio_total += ratio

            # Lưu product được chọn vào SelectedProductForMPS (cho trang MPS)
            try:
                # Tìm product theo source_id (product_code)
                product = Product.objects.get(source_id=product_code)
                SelectedProductForMPS.objects.get_or_create(product=product)
            except Product.DoesNotExist:
                pass  # Nếu product không tìm thấy, bỏ qua

            if not math.isclose(month_ratio_total, 1.0, abs_tol=0.05):
                month_warnings.append(f'{product_code}: tổng ratio = {month_ratio_total:.2f}')

        if month_warnings:
            messages.warning(
                request,
                'Đã lưu kế hoạch phân rã, nhưng một số tháng chưa có tổng ratio bằng 1.00: ' + '; '.join(month_warnings)
            )
        else:
            messages.success(request, 'Lưu kế hoạch phân rã thành công')

        return redirect('product-decomposition')

    plan_rows = []
    allocated_totals = {month.strftime('%Y_%m'): 0.0 for month in months}



    for row_index, row_source in enumerate(existing_product_codes):
        cells = []
        for month in months:
            month_key = month.strftime('%Y_%m')
            ratio_key = f"{row_source['product_code']}_{month_key}"
            ratio_value = float(ratio_map.get(ratio_key) or 0.0)
            qty_value = round(forecast_map[month_key] * ratio_value, 2)

            allocated_totals[month_key] += qty_value
            cells.append({
                'month_key': month_key,
                'ratio': ratio_value,
                'ratio_display': f'{ratio_value:.2f}',
                'qty': qty_value,
            })

        plan_rows.append({
            'row_index': row_index,
            'product_code': row_source['product_code'],
            'product_name': row_source['product_name'],
            'cells': cells,
        })

    if existing_product_codes:
        blank_row_index = len(plan_rows)
        plan_rows.append({
            'row_index': blank_row_index,
            'product_code': '',
            'product_name': '',
            'cells': [
                {
                    'month_key': month.strftime('%Y_%m'),
                    'ratio': 0.0,
                    'qty': 0.0,
                }
                for month in months
            ],
        })

    if not plan_rows:
        for row_index in range(8):
            plan_rows.append({
                'row_index': row_index,
                'product_code': '',
                'product_name': '',
                'cells': [
                    {'month_key': month.strftime('%Y_%m'), 'ratio': 0.0, 'qty': 0.0}
                    for month in months
                ],
            })

    summary_rows = []
    for month in months:
        month_key = month.strftime('%Y_%m')
        total_forecast = forecast_map[month_key]
        allocated_total = round(allocated_totals[month_key], 2)
        summary_rows.append({
            'month': month,
            'month_key': month_key,
            'total_forecast': total_forecast,
            'allocated_total': allocated_total,
            'gap': round(total_forecast - allocated_total, 2),
            'gap_is_zero': math.isclose(total_forecast, allocated_total, abs_tol=0.5),
        })

    plan_data = disaggregate_forecast(
        forecast_list=forecast_list,
        start_month=start_month,
        ratio_qs=ratio_qs,
    )

    forecast_total = round(sum(float(value or 0) for value in forecast_list), 2)

    context = {
        'forecast_mean': forecast_mean,
        'forecast_std': forecast_std,
        'forecast_list': forecast_list,
        'forecast_total': forecast_total,
        'mae': mae,
        'rmse': rmse,
        'mape': mape,
        'months': months,
        'plan_rows': plan_rows,
        'summary_rows': summary_rows,
        'forecast_map': forecast_map,
        'ratio_map': ratio_map,
        'plan_data': plan_data,
        'saved_product_count': len(existing_product_codes),
    }
    return render(request, 'inventory/disaggregate_plan.html', context)


@role_required(ROLE_ADMIN, ROLE_MANAGER, ROLE_STAFF)
def mps(request):
    """Trang MPS (Master Production Schedule)"""
    # Planning dùng mã sản phẩm riêng từ ProductRatio, không dùng Product.id của Operations
    products = (
        ProductRatio.objects
        .values('product_code', 'product_name')
        .order_by('product_code')
        .distinct()
    )
    context = {
        'title': 'MPS - Lập kế hoạch sản xuất chính',
        'products': products,
    }
    return render(request, 'inventory/mps.html', context)


@role_required(ROLE_ADMIN, ROLE_MANAGER, ROLE_STAFF)
def run_mps_api(request):
    """API để tính toán MPS dựa trên PPA"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        import json
        data = json.loads(request.body)
        
        product_code = str(data.get('product_code') or '').strip()
        C = float(data.get('C', 40000000))  # Chi phí thiết lập
        H = float(data.get('H', 250))  # Chi phí lưu kho
        begin_inventory = float(data.get('begin_inventory', 0))
        user_orders = data.get('orders') or []
        
        # Lấy dữ liệu cùng metadata tháng để render đúng timeline
        month_labels = []
        input_labels = []

        if product_code and not product_code.isdigit():
            ratio_rows = list(ProductRatio.objects.filter(product_code=product_code).order_by('month'))
            demand = [int(round(row.forecast_qty or 0)) for row in ratio_rows]
            month_labels = [row.month.strftime('%m/%Y') for row in ratio_rows]

            if ratio_rows:
                start_month = pd.Timestamp(ratio_rows[0].month).to_period('M').to_timestamp()
                input_labels = [
                    (start_month + pd.DateOffset(months=index)).strftime('%m/%Y')
                    for index in range(12)
                ]
        else:
            demand = get_demand_by_product(product_code)

        orders = []
        for index in range(len(demand)):
            if index < len(user_orders):
                try:
                    orders.append(max(0, int(user_orders[index] or 0)))
                except (TypeError, ValueError):
                    orders.append(0)
            else:
                orders.append(0)
        
        if not demand:
            return JsonResponse({
                'error': 'Không có dữ liệu nhu cầu cho sản phẩm này'
            }, status=400)
        
        # Tính toán EPP
        from inventory.services import calculate_epp, calculate_ppa_analysis
        epp = calculate_epp(C, H)
        
        # Tính chi tiết PPA analysis và kích cỡ lô (trả về cả các bước chi tiết)
        ppa_details, lots, ppa_steps = calculate_ppa_analysis(demand, C, H)
        
        # Tính MPS
        projected, atp, net_inventory = calculate_mps(demand, orders, lots, begin_inventory)
        
        # Chuẩn bị dữ liệu để return theo đúng số kỳ forecast hiện có
        if not month_labels:
            months = list(range(1, len(demand) + 1))
            month_labels = [f"Tháng {month}" for month in months]
        else:
            months = list(range(1, len(demand) + 1))
        
        result_data = []
        for i, month in enumerate(months):
            result_data.append({
                'month': month,
                'month_label': month_labels[i] if i < len(month_labels) else f"Tháng {month}",
                'demand': demand[i] if i < len(demand) else 0,
                'orders': orders[i] if i < len(orders) else 0,
                'net_inventory': int(net_inventory[i]) if i < len(net_inventory) else 0,
                'lot_size': lots[i] if i < len(lots) else 0,
                'mps': lots[i] if i < len(lots) else 0,
                'projected_on_hand': int(projected[i]) if i < len(projected) else 0,
                'atp': int(atp[i]) if i < len(atp) else 0,
            })
        
        return JsonResponse({
            'success': True,
            'epp': round(epp, 2),
            'data': result_data,
            'ppa_details': ppa_details,
            'ppa_steps': ppa_steps,
            'month_labels': month_labels,
            'input_labels': input_labels if input_labels else month_labels,
            'demand': demand,
            'orders': orders,
            'mps': lots,
            'projected_on_hand': projected,
            'atp': atp
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'error': f'Lỗi: {str(e)}'
        }, status=500)


@role_required(ROLE_ADMIN, ROLE_MANAGER, ROLE_STAFF)
def mrp(request):
    return render(request, 'inventory/placeholder.html', {
        'title': 'MRP',
        'message': 'Trang MRP đang được xây dựng.'
    })


def _apply_recommendation_filters(recommendations, urgency='ALL', abc='ALL', action='ALL'):
    filtered = recommendations

    if urgency != 'ALL':
        filtered = [item for item in filtered if item.get('urgency') == urgency]

    if abc != 'ALL':
        filtered = [item for item in filtered if item.get('abc_class') == abc]

    if action != 'ALL':
        filtered = [item for item in filtered if item.get('action') == action]
    return filtered


def csrf_failure(request, reason=""):
    message = "Phiên làm việc đã hết hạn hoặc token bảo mật không hợp lệ. Vui lòng tải lại trang và thử lại."
    return render(request, 'inventory/login.html', {
        'error_message': message,
    }, status=403)


session_name_required = role_required(*ALL_ROLE_CODES)
admin_required = role_required(ROLE_ADMIN)


def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    User = get_user_model()
    if not User.objects.filter(username='admin').exists():
        User.objects.create_superuser(username='admin', password='123')

    error_message = None

    if request.method == 'POST':
        username = request.POST.get('display_name', '').strip()
        password = request.POST.get('display_password', '')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            request.session['display_name'] = user.username
            request.session['is_admin'] = user.is_superuser
            request.session['user_role'] = get_user_role_code(user)
            return redirect('dashboard')

        error_message = "Tài khoản hoặc mật khẩu không chính xác."

    return render(request, 'inventory/login.html', {
        'error_message': error_message,
    })


def logout_view(request):
    logout(request)
    return redirect('login')

# =========================
# DASHBOARD (SALES + TRANSACTION)
# =========================
@session_name_required
def dashboard(request):

    # =========================
    # POST HANDLE
    # =========================
    if request.method == 'POST':

        # ===== ADD SALES =====
        if 'add_sales' in request.POST:
            product_id = request.POST.get('product_id')
            quantity = request.POST.get('quantity')
            date_input = request.POST.get('date')

            # ép kiểu quantity
            try:
                quantity = int(quantity)
            except:
                quantity = 0

            # ép kiểu date
            try:
                date_input = datetime.strptime(date_input, "%Y-%m-%d").date()
            except:
                date_input = None

            if product_id and quantity > 0 and date_input:
                SalesData.objects.create(
                    product_id=product_id,
                    quantity=quantity,
                    date=date_input,
                    source=SalesData.SOURCE_OPERATIONS,
                )

        # ===== ADD TRANSACTION =====
        elif 'add_transaction' in request.POST:
            material_id = request.POST.get('material_id')
            quantity = request.POST.get('quantity')
            t_type = request.POST.get('transaction_type')
            date_input = request.POST.get('date')

            try:
                quantity = int(quantity)
            except:
                quantity = 0

            try:
                date_input = datetime.strptime(date_input, "%Y-%m-%d").date()
            except:
                date_input = None

            if material_id and quantity > 0 and date_input:
                material = Material.objects.get(id=material_id)

                # update tồn kho
                if t_type == 'IN':
                    material.on_hand += quantity
                elif t_type == 'OUT':
                    if material.on_hand >= quantity:
                        material.on_hand -= quantity

                material.save()

                # lưu transaction
                Transaction.objects.create(
                    material=material,
                    quantity=quantity,
                    transaction_type=t_type,
                    date=date_input
                )

    # =========================
    # GET DATA
    # =========================
    products = Product.objects.all()
    materials = Material.objects.all()

    # 👉 lấy mới nhất + ổn định thứ tự
    sales_qs = SalesData.objects.filter(source=SalesData.SOURCE_OPERATIONS)
    sales_list = sales_qs.order_by('-date', '-id')[:10]
    transaction_list = Transaction.objects.filter(source=Transaction.SOURCE_OPERATIONS).order_by('-id')[:10]

    # =========================
    # TOP 5 PRODUCTS BY REVENUE (unit_price derived from BOM materials)
    # revenue = unit_price * total_quantity_sold
    # =========================
    sales_totals = {}
    for sale in sales_qs.select_related('product').all():
        product_id = sale.product_id
        product_name = sale.product.name if sale.product else str(product_id)
        bucket = sales_totals.setdefault(
            product_id,
            {'product_id': product_id, 'product__name': product_name, 'total_quantity': 0},
        )
        bucket['total_quantity'] += float(sale.quantity or 0)

    sales_agg = list(sales_totals.values())

    product_revenues = []
    total_revenue = 0.0

    for row in sales_agg:
        product_id = row.get('product_id')
        product_name = row.get('product__name')
        qty = float(row.get('total_quantity') or 0)

        # compute unit price from BOM: sum(material.price_cost * quantity_per_unit)
        boms = BOM.objects.filter(product_id=product_id).select_related('material')
        unit_price = 0.0
        for bom in boms:
            if bom.material and bom.material.price_cost:
                unit_price += float(bom.quantity_per_unit or 0) * float(bom.material.price_cost or 0)

        revenue = qty * unit_price
        total_revenue += revenue

        product_revenues.append({
            'product_id': product_id,
            'name': product_name,
            'quantity': qty,
            'unit_price': unit_price,
            'revenue': revenue,
        })

    # sort by revenue desc and take top 5
    product_revenues.sort(key=lambda x: x['revenue'], reverse=True)
    top5 = product_revenues[:5]

    top_sales_chart_labels = []
    top_sales_chart_values = []
    top_sales_chart_cumulative_percent = []

    for item in top5:
        # percent of total revenue for this product (not cumulative)
        percent = (item['revenue'] / total_revenue * 100) if total_revenue else 0

        top_sales_chart_labels.append(item['name'])
        top_sales_chart_values.append(round(item['revenue'], 2))
        top_sales_chart_cumulative_percent.append(round(percent, 1))

    top_sales_chart_data = {
        'labels': top_sales_chart_labels,
        'values': top_sales_chart_values,
        'cumulative_percent': top_sales_chart_cumulative_percent,
        'total_revenue': round(float(total_revenue), 2),
    }

    # 👉 Gợi ý ngày nhập gần nhất
    last_sale = sales_qs.order_by('-date').first()

    if last_sale:
        min_date = last_sale.date.strftime("%Y-%m-%d")
    else:
        min_date = None

    today = date.today().strftime("%Y-%m-%d")

    # =========================
    # ABC VALUE SUMMARY (TRANSACTION OUT)
    # =========================
    demand_map = {}
    for transaction in Transaction.objects.filter(transaction_type='OUT', source=Transaction.SOURCE_OPERATIONS).select_related('material').all():
        demand_map[transaction.material_id] = demand_map.get(transaction.material_id, 0) + float(transaction.quantity or 0)

    abc_material_list = []
    for material in materials:
        abc_material_list.append({
            'material': material,
            'demand': demand_map.get(material.id, 0),
        })

    abc_map = abc_classification(abc_material_list)
    abc_value_total = {'A': 0.0, 'B': 0.0, 'C': 0.0}
    abc_counts = {'A': 0, 'B': 0, 'C': 0}

    for item in abc_material_list:
        material = item['material']
        demand = item['demand'] or 0
        abc_class = abc_map.get(material.id)

        if abc_class:
            abc_value_total[abc_class] += float(demand) * float(material.price_cost)
            # count materials per ABC class (for pie chart proportions)
            if abc_class in abc_counts:
                abc_counts[abc_class] += 1

    abc_chart_data = {
        'labels': ['A', 'B', 'C'],
        'values': [
            round(abc_value_total['A'], 2),
            round(abc_value_total['B'], 2),
            round(abc_value_total['C'], 2),
        ],
        'colors': ['#4b5f70', '#4e8b90', '#4f8e74'],
        'total': round(sum(abc_value_total.values()), 2),
    }

    selected_urgency = request.GET.get('urgency', 'ALL').upper()
    selected_abc = request.GET.get('abc', 'ALL').upper()
    selected_action = request.GET.get('action', 'ALL').upper()

    try:
        recommendations_all, recommendation_summary = build_dashboard_recommendations(limit=200, source=SalesData.SOURCE_OPERATIONS)
    except Exception:
        recommendations_all, recommendation_summary = [], {
            'urgent': 0,
            'medium': 0,
            'low': 0,
            'order': 0,
            'safe': materials.count(),
        }
    recommendations_filtered_all = recommendations_all
    recommendations = recommendations_filtered_all
    recommendation_display = f"{len(recommendations)} URGENT"

    return render(request, 'inventory/dashboard.html', {
        'products': products,
        'materials': materials,

        'total_products': products.count(),
        'total_materials': materials.count(),

        'sales_list': sales_list,
        'transaction_list': transaction_list,
        'top_sales_chart_data': top_sales_chart_data,
        'recommendations': recommendations,
        'recommendation_summary': recommendation_summary,
        'recommendation_total': len(recommendations_filtered_all),
        'recommendation_filtered': len(recommendations),
        'recommendation_display': recommendation_display,
        'min_date': min_date,
        'today': today,
        'abc_value_total': abc_value_total,
        'abc_value_sum': sum(abc_value_total.values()),
        'abc_counts': abc_counts,
        'abc_count_total': sum(abc_counts.values()),
        'abc_chart_data': abc_chart_data,
    })


@session_name_required
def recommendation_api(request):
    selected_urgency = request.GET.get('urgency', 'ALL').upper()
    selected_abc = request.GET.get('abc', 'ALL').upper()
    selected_action = request.GET.get('action', 'ALL').upper()

    try:
        recommendations, summary = build_dashboard_recommendations(limit=200)
    except Exception:
        recommendations, summary = [], {
            'urgent': 0,
            'medium': 0,
            'low': 0,
            'order': 0,
            'safe': Material.objects.count(),
        }
    filtered = _apply_recommendation_filters(
        recommendations,
        urgency=selected_urgency,
        abc=selected_abc,
        action=selected_action,
    )
    payload = [item['json'] for item in filtered]

    return JsonResponse({
        'summary': summary,
        'filters': {
            'urgency': selected_urgency,
            'abc': selected_abc,
            'action': selected_action,
            'count': len(payload),
        },
        'recommendations': payload,
    }, json_dumps_params={'ensure_ascii': False, 'indent': 2})


@session_name_required
def recommendation_excel_api(request):
    selected_urgency = request.GET.get('urgency', 'ALL').upper()
    selected_abc = request.GET.get('abc', 'ALL').upper()
    selected_action = request.GET.get('action', 'ALL').upper()

    recommendations, _ = build_dashboard_recommendations(limit=500)
    filtered = _apply_recommendation_filters(
        recommendations,
        urgency=selected_urgency,
        abc=selected_abc,
        action=selected_action,
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Recommendations'

    headers = [
        'item',
        'material_name',
        'abc_class',
        'action',
        'quantity',
        'urgency',
        'ip',
        'rop',
        's',
        'message',
    ]
    worksheet.append(headers)

    for item in filtered:
        worksheet.append([
            item.get('item'),
            item.get('material_name'),
            item.get('abc_class'),
            item.get('action'),
            item.get('quantity'),
            item.get('urgency'),
            item.get('ip'),
            item.get('rop'),
            item.get('s'),
            item.get('message'),
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="dashboard_recommendations.xlsx"'
    workbook.save(response)

    return response


# =========================
# DELETE SALE
# =========================
@session_name_required
def delete_sale(request, id):
    sale = get_object_or_404(SalesData, id=id)
    sale.delete()
    return redirect('dashboard')


@session_name_required
def delete_transaction(request, id):
    t = get_object_or_404(Transaction, id=id)
    t.delete()
    return redirect('dashboard')


@role_required(ROLE_ADMIN, ROLE_MANAGER, ROLE_STAFF)
def transaction_list(request):
    transactions = Transaction.objects.select_related('material').order_by('-date', '-id')
    return render(request, 'inventory/transaction_list.html', {
        'transactions': transactions,
    })


@role_required(ROLE_ADMIN, ROLE_MANAGER, ROLE_STAFF)
def transaction_create(request):
    materials = Material.objects.all().order_by('name')

    if request.method == 'POST':
        form = TransactionForm(request.POST)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.source = Transaction.SOURCE_OPERATIONS
            transaction.save()
            return redirect('transaction-list')
    else:
        form = TransactionForm()

    return render(request, 'inventory/transaction_form.html', {
        'form': form,
        'materials': materials,
    })


# =========================
# PRODUCT + DSS
# =========================
@role_required(ROLE_ADMIN, ROLE_MANAGER, ROLE_STAFF)
def product_list(request):
    from .models import SalesData
    
    products = Product.objects.all()

    results = []
    selected_product = None

    if request.method == 'POST':
        product_id = request.POST.get('product_id')

        if product_id:
            selected_product = Product.objects.get(id=product_id)
            # Use OPERATIONS source data
            results = run_dss(product_id, source=SalesData.SOURCE_OPERATIONS)

    return render(request, 'inventory/product_list.html', {
        'products': products,
        'results': results,
        'selected_product': selected_product,
    })


# =========================
# MATERIAL LIST
# =========================
@role_required(ROLE_ADMIN, ROLE_MANAGER)
def material_list(request):
    materials = Material.objects.all()

    data = []
    for m in materials:
        data.append({
            'material_id': m.source_id or f'NVL{m.id}',
            'name': m.name,
            'on_hand': m.on_hand,
            'on_order': m.on_order,
            'inventory_position': m.on_hand + m.on_order,
            'leadtime': m.leadtime,
            'holding_cost': m.holding_cost,
            'ordering_cost': m.ordering_cost,
            'price_cost': m.price_cost,
        })

    return render(request, 'inventory/material_list.html', {
        'materials': data
    })


# =========================
# OTHER PAGES
# =========================
@role_required(ROLE_ADMIN, ROLE_MANAGER, ROLE_STAFF)
def alert(request):
    alerts, summary = build_inventory_alert_recommendations(source=SalesData.SOURCE_OPERATIONS)
    watchlist = build_inventory_watchlist_recommendations(source=SalesData.SOURCE_OPERATIONS)

    return render(request, 'inventory/alert.html', {
        "alerts": alerts,
        "summary": summary,
        "watchlist": watchlist,
    })


@role_required(ROLE_ADMIN, ROLE_MANAGER, ROLE_STAFF)
def forecast(request):
    from .models import Product, BOM
    from .services import forecast_product
    from collections import defaultdict
    import math

    products = Product.objects.all()

    selected_product = None
    product_result = None
    material_results = []
    forecast_7 = []

    # OPERATIONS forecast uses operations source only.
    sales_qs = SalesData.objects.filter(source=SalesData.SOURCE_OPERATIONS)

    # =========================
    # PRODUCT FORECAST
    # =========================
    if request.method == 'POST':
        product_id = request.POST.get('product_id')

        if product_id:
            selected_product = Product.objects.get(id=product_id)

            mean, std, forecast_7, mae, rmse, mape = forecast_product(product_id, sales_qs=sales_qs)

            mape_value = float(round(mape, 2))
            if mape_value < 10:
                mape_level = "good"
                evaluation_class = "evaluation-good"
                evaluation_icon = "fas fa-check"
                evaluation_message = (
                    "Mô hình có độ chính xác rất cao (MAPE < 10%). "
                    "Kết quả dự báo rất đáng tin cậy."
                )
            elif mape_value < 20:
                mape_level = "medium"
                evaluation_class = "evaluation-medium"
                evaluation_icon = "fas fa-info"
                evaluation_message = (
                    "Mô hình ở mức chấp nhận được (10% - 20%). "
                    "Cần theo dõi thêm biến động thực tế."
                )
            else:
                mape_level = "bad"
                evaluation_class = "evaluation-bad"
                evaluation_icon = "fas fa-exclamation-triangle"
                evaluation_message = (
                    "Độ sai số cao (MAPE > 20%). Nhu cầu sản phẩm này có biến động "
                    "quá lớn, mô hình hiện tại không phù hợp."
                )

            product_result = {
                "name": selected_product.name,
                "mean": round(mean, 2),
                "std": round(std, 2),
                "forecast_7": [round(x, 2) for x in forecast_7],
                "mae": round(mae, 2),
                "rmse": round(rmse, 2),
                "mape": mape_value,
                "mape_level": mape_level,
                "evaluation_class": evaluation_class,
                "evaluation_icon": evaluation_icon,
                "evaluation_message": evaluation_message,
            }

    # =========================
    # 🔥 MATERIAL AGGREGATE (LUÔN CHẠY)
    # =========================
    material_dict = defaultdict(lambda: {
        "material": "",
        "material_id": "",
        "mean": 0,
        "variance": 0,
        "forecast_7": [0] * 7
    })

    # ✅ FIX: Fetch all BOMs with material eagerly (NOT nested loop)
    all_boms = BOM.objects.select_related('material', 'product')
    
    # 🔥 FIX: Use safe defaults - SKIP expensive forecasting on page load.
    # When user requests analysis, build forecasts only for the selected
    # product and other products that share BOM materials with it.
    product_forecasts = {}
    if request.method == 'POST' and selected_product:
        try:
            if product_result is not None:
                product_forecasts[selected_product.id] = (
                    float(product_result.get("mean", 0) or 0),
                    float(product_result.get("std", 0) or 0),
                    [float(x or 0) for x in (forecast_7 or [0] * 7)],
                )
            else:
                mean, std, product_forecast_7, _, _, _ = forecast_product(selected_product.id, sales_qs=sales_qs)
                product_forecasts[selected_product.id] = (mean, std, product_forecast_7)

            selected_material_ids = set(
                BOM.objects.filter(product=selected_product).values_list('material_id', flat=True)
            )

            shared_product_ids = set()
            if selected_material_ids:
                shared_product_ids = set(
                    BOM.objects.filter(material_id__in=selected_material_ids)
                    .values_list('product_id', flat=True)
                    .distinct()
                )

            for product_id in shared_product_ids:
                if product_id in product_forecasts:
                    continue

                try:
                    mean, std, product_forecast_7, _, _, _ = forecast_product(product_id, sales_qs=sales_qs)
                    product_forecasts[product_id] = (mean, std, product_forecast_7)
                except Exception as e:
                    print(f"Forecast error for shared product {product_id}: {e}")
        except Exception as e:
            print(f"Forecast error: {e}")
    
    # Use safe defaults for material aggregation (no forecasting on page load)
    selected_boms = []
    shared_boms = []
    if selected_product:
        selected_boms = list(
            BOM.objects.filter(product=selected_product).select_related('material')
        )
        selected_material_ids = set(bom.material_id for bom in selected_boms)
        shared_boms = list(
            BOM.objects.filter(material_id__in=selected_material_ids)
            .exclude(product=selected_product)
            .select_related('material', 'product')
        )
    else:
        selected_material_ids = set()

    # Aggregate duplicate BOM rows by material/product so the view is stable even if
    # the database contains repeated BOM entries.
    # For each material, keep the FIRST quantity (don't sum duplicates).
    selected_quantities = {}
    selected_materials = {}
    for bom in selected_boms:
        if bom.material_id not in selected_quantities:
            selected_quantities[bom.material_id] = bom.quantity_per_unit or 0
            selected_materials[bom.material_id] = bom.material

    shared_quantities_by_material = {}
    shared_products_by_material = {}
    for bom in shared_boms:
        mat_id = bom.material_id
        prod_id = bom.product_id
        qty = bom.quantity_per_unit or 0
        if mat_id not in shared_quantities_by_material:
            shared_quantities_by_material[mat_id] = {}
        if prod_id not in shared_quantities_by_material[mat_id]:
            shared_quantities_by_material[mat_id][prod_id] = qty
            shared_products_by_material.setdefault(mat_id, {})
            shared_products_by_material[mat_id][prod_id] = bom.product

    material_results = []
    if selected_product:
        selected_stats = product_forecasts.get(selected_product.id, (0.0, 0.0, [0] * 7))
        selected_mean = float(selected_stats[0] or 0.0)
        selected_std = float(selected_stats[1] or 0.0)
        selected_forecast_7 = [float(x or 0.0) for x in selected_stats[2]]

        material_shared_boms = {}
        for bom in shared_boms:
            material_shared_boms.setdefault(bom.material_id, []).append(bom)

        for material_id, quantity in selected_quantities.items():
            material = selected_materials.get(material_id)
            qty = float(quantity or 0)
            material_mean = selected_mean * qty
            material_variance = (selected_std * qty) ** 2
            material_forecast_7 = [round(val * qty, 2) for val in selected_forecast_7]

            for bom in material_shared_boms.get(material_id, []):
                other_stats = product_forecasts.get(bom.product_id, (0.0, 0.0, [0] * 7))
                other_mean = float(other_stats[0] or 0.0)
                other_std = float(other_stats[1] or 0.0)
                other_forecast_7 = [float(x or 0.0) for x in other_stats[2]]
                other_qty = float(bom.quantity_per_unit or 0)

                material_mean += other_mean * other_qty
                material_variance += (other_std * other_qty) ** 2
                material_forecast_7 = [
                    round(current + other_val * other_qty, 2)
                    for current, other_val in zip(material_forecast_7, other_forecast_7)
                ]

            material_results.append({
                "material": material.name,
                "material_id": material.source_id or f"NVL{material.id}",
                "mean": round(material_mean, 2),
                "std": round(math.sqrt(material_variance), 2),
                "forecast_7": material_forecast_7,
            })

    # =========================
    # 🔥 LUÔN RETURN
    # =========================
    return render(request, 'inventory/forecast.html', {
        "products": products,
        "selected_product": selected_product,
        "product_result": product_result,
        "material_results": material_results,
        "forecast_7": forecast_7
    })


@role_required(ROLE_ADMIN, ROLE_MANAGER, ROLE_STAFF)
def forecast_monthly(request):
    from collections import defaultdict
    from django.utils import timezone
    from .models import MonthlyProductionData

    form = MonthlyForecastImportForm()
    result_message = None
    error_message = None

    def _parse_month_value(raw_value):
        if hasattr(raw_value, 'date') and callable(raw_value.date):
            raw_value = raw_value.date()
        if isinstance(raw_value, date):
            return raw_value.replace(day=1)
        if isinstance(raw_value, str):
            text = raw_value.strip()
            for fmt in ('%m/%Y', '%m-%Y', '%Y-%m', '%Y/%m'):
                try:
                    parsed = datetime.strptime(text, fmt)
                    return parsed.date().replace(day=1)
                except ValueError:
                    continue
        return None

    if request.method == 'POST' and 'import_monthly_excel' in request.POST:
        form = MonthlyForecastImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']
            try:
                wb = load_workbook(excel_file)
                ws = wb.active

                rows_imported = 0
                rows_skipped = 0
                imported_rows = []
                aggregated = {}

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or row[0] is None:
                        continue

                    month_obj = _parse_month_value(row[0])
                    try:
                        quantity = float(str(row[1]).replace(',', '').strip())
                    except Exception:
                        continue

                    if not month_obj:
                        continue

                    aggregated[month_obj] = aggregated.get(month_obj, 0) + quantity

                if aggregated:
                    MonthlyProductionData.objects.filter(source=MonthlyProductionData.SOURCE_PLANNING).delete()
                    for month_obj in sorted(aggregated.keys()):
                        quantity = float(aggregated[month_obj])
                        MonthlyProductionData.objects.create(
                            month=month_obj,
                            quantity=quantity,
                            source=MonthlyProductionData.SOURCE_PLANNING,
                        )
                        imported_rows.append({
                            'month': month_obj,
                            'quantity': quantity,
                        })
                        rows_imported += 1

                result_message = f'Đã import {rows_imported} dòng dữ liệu sản xuất quá khứ thành công.'
                if rows_skipped:
                    result_message += f' Bỏ qua {rows_skipped} dòng không hợp lệ.'
            except Exception as e:
                error_message = f'Lỗi xử lý file: {str(e)}'
        else:
            error_message = 'Vui lòng chọn file Excel hợp lệ.'

    history_qs = MonthlyProductionData.objects.filter(source=MonthlyProductionData.SOURCE_PLANNING).order_by('month')
    history_rows = [
        {
            'month': item.month.strftime('%m/%Y'),
            'quantity': float(item.quantity or 0),
        }
        for item in history_qs
    ]

    forecast_mean, forecast_std, forecast_8, mae, rmse, mape = forecast_monthly_total(history_qs=history_qs)
    forecast_result = None
    if history_rows:
        forecast_result = {
            'mean': round(forecast_mean, 2),
            'std': round(forecast_std, 2),
            'forecast_8': [round(x, 2) for x in forecast_8],
            'mae': round(mae, 2),
            'rmse': round(rmse, 2),
            'mape': round(mape, 2),
        }

    forecast_8_rows = []
    if history_qs.exists():
        last_month = history_qs.last().month
        for idx, value in enumerate(forecast_8, start=1):
            month_label = (last_month.replace(day=1) + timedelta(days=32 * idx)).replace(day=1)
            forecast_8_rows.append({
                'month': month_label.strftime('%m/%Y'),
                'quantity': round(float(value or 0), 2),
            })

    return render(request, 'inventory/forecast_monthly_import.html', {
        'form': form,
        'result_message': result_message,
        'error_message': error_message,
        'history_rows': history_rows,
        'forecast_result': forecast_result,
        'forecast_8_rows': forecast_8_rows,
    })
from .services import abc_classification, forecast_product
from .models import Product, BOM

@role_required(ROLE_ADMIN, ROLE_MANAGER, ROLE_STAFF)
def abc_page(request):
    from .models import Product, Material, Transaction, BOM
    from django.db.models import Sum
    from .services import abc_classification

    products = Product.objects.all()

    product_id = request.GET.get("product_id")
    abc_filter = request.GET.get("abc")

    selected_product = None
    results = []

    # =========================
    # 🔵 1. ABC TOÀN HỆ THỐNG (THEO TRANSACTION - TỐI ƯU)
    # =========================
    material_list = []

    # 👉 gom transaction 1 lần (KHÔNG loop)
    try:
        transaction_data = (
            Transaction.objects
            .filter(transaction_type='OUT', source=Transaction.SOURCE_OPERATIONS)
            .values('material')
            .annotate(total=Sum('quantity'))
        )
        # 👉 convert thành dict cho nhanh
        demand_map = {
            item['material']: item['total']
            for item in transaction_data
        }
    except Exception as e:
        # Fallback if aggregation fails (Djongo issue)
        print(f"Transaction aggregation error: {e}")
        demand_map = {}

    materials = Material.objects.all()

    for m in materials:
        demand = demand_map.get(m.id, 0)

        material_list.append({
            "material": m,
            "demand": demand  # 🔥 giữ tên mean cho abc_classification
        })

    # 👉 ABC
    abc_all = abc_classification(material_list)

    # 👉 đếm tổng
    abc_total = {"A": 0, "B": 0, "C": 0}

    for item in material_list:
        m = item["material"]
        cat = abc_all.get(m.id)

        if cat:
            abc_total[cat] += 1

    # =========================
    # 🔴 FILTER MATERIAL
    # =========================
    filtered_materials = []

    for item in material_list:
        m = item["material"]
        demand = item["demand"]
        abc = abc_all.get(m.id, "-")

        if not abc_filter or abc == abc_filter:
            filtered_materials.append({
                "material_id": m.source_id or f"NVL{m.id}",
                "material": m.name,
                "demand": demand,
                "price": m.price_cost,
                "value": demand * m.price_cost,
                "abc": abc
            })

    # =========================
    # 🟢 ABC THEO PRODUCT
    # =========================
    if product_id:
        selected_product = Product.objects.get(id=product_id)

        boms = BOM.objects.filter(product=selected_product).select_related('material')

        for bom in boms:
            m = bom.material
            demand = demand_map.get(m.id, 0)

            results.append({
                "material_id": m.source_id or f"NVL{m.id}",
                "material": m.name,
                "demand": demand,
                "price": m.price_cost,
                "value": demand * m.price_cost,
                "abc": abc_all.get(m.id, "-")
            })

    return render(request, "inventory/abc.html", {
        "products": products,
        "selected_product": selected_product,
        "results": results,
        "filtered_materials": filtered_materials,
        "abc_total": abc_total,
        "selected_abc": abc_filter
    })


@role_required(ROLE_ADMIN, ROLE_MANAGER, ROLE_STAFF)
def inventory_analysis(request):
    from .models import Material, Transaction
    from django.db.models import Sum, Q
    
    materials = Material.objects.all()
    
    # Get stock-out data (OPERATIONS source only)
    try:
        transaction_data = (
            Transaction.objects
            .filter(transaction_type='OUT', source=Transaction.SOURCE_OPERATIONS)
            .values('material')
            .annotate(total_out=Sum('quantity'))
        )
        stock_out_map = {
            item['material']: item['total_out']
            for item in transaction_data
        }
    except:
        stock_out_map = {}
    
    # Build analysis data
    analysis_data = []
    
    for material in materials:
        total_out = stock_out_map.get(material.id, 0)
        
        # Determine status based on on_hand level
        if material.on_hand <= 0:
            status = 'out_of_stock'
            status_label = 'Hết hàng'
            status_color = 'danger'
        elif material.on_hand < material.leadtime * 10:  # Simple threshold
            status = 'low_stock'
            status_label = 'Tồn kho thấp'
            status_color = 'warning'
        else:
            status = 'normal'
            status_label = 'Bình thường'
            status_color = 'success'
        
        analysis_data.append({
            'material_id': material.source_id or f'NVL{material.id}',
            'material': material,
            'on_hand': material.on_hand,
            'on_order': material.on_order,
            'total_demand': total_out,
            'leadtime': material.leadtime,
            'holding_cost': material.holding_cost,
            'ordering_cost': material.ordering_cost,
            'price': material.price_cost,
            'value': material.on_hand * material.price_cost,
            'status': status,
            'status_label': status_label,
            'status_color': status_color
        })
    
    # Filter options
    status_filter = request.GET.get('status', 'all')
    if status_filter != 'all':
        analysis_data = [item for item in analysis_data if item['status'] == status_filter]
    
    # Sort options
    sort_by = request.GET.get('sort', 'material')
    if sort_by == 'value':
        analysis_data.sort(key=lambda x: x['value'], reverse=True)
    elif sort_by == 'on_hand':
        analysis_data.sort(key=lambda x: x['on_hand'], reverse=True)
    elif sort_by == 'demand':
        analysis_data.sort(key=lambda x: x['total_demand'], reverse=True)
    else:
        analysis_data.sort(key=lambda x: x['material'].name)
    
    # Calculate summary stats
    total_inventory_value = sum(item['value'] for item in analysis_data)
    total_on_hand = sum(item['on_hand'] for item in analysis_data)
    
    return render(request, 'inventory/inventory_analysis.html', {
        'analysis_data': analysis_data,
        'status_filter': status_filter,
        'sort_by': sort_by,
        'total_inventory_value': total_inventory_value,
        'total_on_hand': total_on_hand,
    })


@admin_required
def system_settings(request):
    return render(request, "inventory/system_settings.html")


@admin_required
def access_control(request):
    User = get_user_model()
    ensure_role_groups()
    error_message = None
    success_message = None

    if request.method == 'POST':
        role_update_user_id = request.POST.get('update_role_user_id')
        delete_user_id = request.POST.get('delete_user_id')

        if role_update_user_id:
            try:
                target_user = User.objects.get(id=role_update_user_id)
            except User.DoesNotExist:
                error_message = 'Tài khoản cần cập nhật không tồn tại.'
            else:
                if target_user.id == request.user.id:
                    error_message = 'Bạn không thể tự thay đổi vai trò của tài khoản đang đăng nhập.'
                else:
                    selected_role = request.POST.get('role_code', ROLE_OTHER)
                    if selected_role == ROLE_ADMIN:
                        if not target_user.is_superuser:
                            target_user.is_superuser = True
                            target_user.save(update_fields=['is_superuser'])
                        target_user.groups.clear()
                    else:
                        if target_user.is_superuser:
                            target_user.is_superuser = False
                            target_user.save(update_fields=['is_superuser'])
                        assign_user_role(target_user, selected_role)
                    success_message = f'Đã cập nhật vai trò cho {target_user.username}.'

        elif delete_user_id:
            if not request.user.is_superuser:
                error_message = 'Chỉ admin mới có quyền xóa tài khoản.'
            else:
                try:
                    target_user = User.objects.get(id=delete_user_id)
                except User.DoesNotExist:
                    error_message = 'Tài khoản cần xóa không tồn tại.'
                else:
                    if target_user.id == request.user.id:
                        error_message = 'Bạn không thể tự xóa tài khoản đang đăng nhập.'
                    elif target_user.is_superuser and User.objects.filter(is_superuser=True).count() <= 1:
                        error_message = 'Không thể xóa admin cuối cùng của hệ thống.'
                    else:
                        deleted_username = target_user.username
                        target_user.delete()
                        success_message = f'Đã xóa tài khoản {deleted_username}.'
        else:
            username = request.POST.get('username', '').strip()
            password = request.POST.get('password', '')
            role_code = request.POST.get('role_code', ROLE_OTHER)

            if not username or not password:
                error_message = 'Vui lòng nhập đầy đủ tài khoản và mật khẩu.'
            elif User.objects.filter(username=username).exists():
                error_message = 'Tài khoản đã tồn tại, vui lòng chọn tên khác.'
            else:
                create_user_with_role(username=username, password=password, role_code=role_code)
                success_message = f'Đã tạo tài khoản {username} thành công.'

    users = []
    for user in User.objects.all().order_by('username'):
        users.append({
            'id': user.id,
            'username': user.username,
            'is_superuser': user.is_superuser,
            'is_active': user.is_active,
            'role_code': get_user_role_code(user),
            'role_label': get_user_role_label(user),
        })

    return render(request, 'inventory/access_control.html', {
        'users': users,
        'error_message': error_message,
        'success_message': success_message,
        'role_choices': ROLE_FORM_CHOICES,
        'default_role_code': ROLE_OTHER,
    })


# =========================
# IMPORT DATA FROM EXCEL
# =========================
@role_required(ROLE_ADMIN, ROLE_MANAGER, ROLE_STAFF)
def import_data(request, sales_source=SalesData.SOURCE_OPERATIONS):
    def _parse_excel_date(raw_value):
        if isinstance(raw_value, datetime):
            return raw_value.date()
        if isinstance(raw_value, date):
            return raw_value
        if isinstance(raw_value, str):
            text = raw_value.strip()
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                try:
                    return datetime.strptime(text, fmt).date()
                except ValueError:
                    continue
        return None

    def _parse_transaction_type(raw_value):
        text = str(raw_value).strip().upper()
        if text in ('IN', 'NHAP', 'NHAP KHO'):
            return 'IN'
        if text in ('OUT', 'XUAT', 'XUAT KHO'):
            return 'OUT'
        return None

    if sales_source not in dict(SalesData.SOURCE_CHOICES):
        sales_source = SalesData.SOURCE_OPERATIONS

    source_label = 'OPERATIONS' if sales_source == SalesData.SOURCE_OPERATIONS else 'PLANNING'
    page_title = 'Import dữ liệu' if sales_source == SalesData.SOURCE_OPERATIONS else 'Import dữ liệu kế hoạch'
    page_subtitle = (
        'Nhập dữ liệu bán hàng hoặc giao dịch cho OPERATIONS'
        if sales_source == SalesData.SOURCE_OPERATIONS
        else 'Nhập dữ liệu bán hàng cho PLANNING'
    )

    form = ImportDataForm()
    result_message = None
    error_message = None
    products = Product.objects.all()
    materials = Material.objects.all()

    last_sale = SalesData.objects.filter(source=sales_source).order_by('-date').first()
    min_date = last_sale.date.strftime("%Y-%m-%d") if last_sale else None
    today = date.today().strftime("%Y-%m-%d")

    if request.method == 'POST':
        if 'add_sales' in request.POST:
            product_id = request.POST.get('product_id')
            quantity = request.POST.get('quantity')
            date_input = request.POST.get('date')

            try:
                quantity = int(quantity)
            except Exception:
                quantity = 0

            try:
                date_input = datetime.strptime(date_input, "%Y-%m-%d").date()
            except Exception:
                date_input = None

            if product_id and quantity > 0 and date_input:
                is_duplicate = SalesData.objects.filter(
                    product_id=product_id,
                    quantity=quantity,
                    date=date_input,
                    source=sales_source,
                ).exists()

                if is_duplicate:
                    result_message = 'Bỏ qua: bản ghi doanh số trùng hoàn toàn đã tồn tại.'
                else:
                    SalesData.objects.create(
                        product_id=product_id,
                        quantity=quantity,
                        date=date_input,
                        source=sales_source,
                    )
                    result_message = f'Đã thêm doanh số thành công cho {source_label}.'
            else:
                error_message = 'Dữ liệu nhập doanh số không hợp lệ.'

        elif 'add_transaction' in request.POST:
            material_id = request.POST.get('material_id')
            quantity = request.POST.get('quantity')
            t_type = request.POST.get('transaction_type')
            date_input = request.POST.get('date')

            try:
                quantity = int(quantity)
            except Exception:
                quantity = 0

            try:
                date_input = datetime.strptime(date_input, "%Y-%m-%d").date()
            except Exception:
                date_input = None

            if material_id and quantity > 0 and date_input and t_type in ('IN', 'OUT'):
                material = Material.objects.get(id=material_id)

                is_duplicate = Transaction.objects.filter(
                    material=material,
                    quantity=quantity,
                    transaction_type=t_type,
                    date=date_input,
                ).exists()

                if is_duplicate:
                    result_message = 'Bỏ qua: giao dịch trùng hoàn toàn đã tồn tại.'
                else:
                    if t_type == 'IN':
                        material.on_hand += quantity
                    elif t_type == 'OUT' and material.on_hand >= quantity:
                        material.on_hand -= quantity

                    material.save()

                    Transaction.objects.create(
                        material=material,
                        quantity=quantity,
                        transaction_type=t_type,
                        date=date_input
                    )
                    result_message = 'Đã thêm giao dịch thành công.'
            else:
                error_message = 'Dữ liệu nhập giao dịch không hợp lệ.'

        elif 'import_excel' in request.POST:
            form = ImportDataForm(request.POST, request.FILES)
            if form.is_valid():
                import_type = form.cleaned_data['import_type']
                excel_file = request.FILES['file']

                try:
                    wb = load_workbook(excel_file)
                    ws = wb.active

                    if import_type == 'sales':
                        # Expected columns: product_name, date, quantity
                        rows_imported = 0
                        rows_skipped_duplicate = 0
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if row[0] is None:
                                continue
                            try:
                                product_name = str(row[0]).strip()
                                date_obj = _parse_excel_date(row[1])
                                quantity = int(row[2])

                                if not date_obj:
                                    continue

                                product = Product.objects.filter(name=product_name).first() or \
                                    Product.objects.filter(source_id=product_name).first()

                                if product:
                                    is_duplicate = SalesData.objects.filter(
                                        product=product,
                                        date=date_obj,
                                        quantity=quantity,
                                        source=sales_source,
                                    ).exists()

                                    if is_duplicate:
                                        rows_skipped_duplicate += 1
                                    else:
                                        SalesData.objects.create(
                                            product=product,
                                            date=date_obj,
                                            quantity=quantity,
                                            source=sales_source,
                                        )
                                        rows_imported += 1
                            except Exception:
                                continue

                        result_message = (
                            f'✓ Đã import {rows_imported} bản ghi bán hàng ({source_label}) thành công. '
                            f'Bỏ qua {rows_skipped_duplicate} bản ghi trùng hoàn toàn.'
                        )

                    elif import_type == 'transaction':
                        if sales_source != SalesData.SOURCE_OPERATIONS:
                            error_message = 'Nguồn PLANNING không dùng dữ liệu giao dịch kho.'
                            return render(request, 'inventory/import_data.html', {
                                'form': form,
                                'result_message': result_message,
                                'error_message': error_message,
                                'products': products,
                                'materials': materials,
                                'min_date': min_date,
                                'today': today,
                                'sales_source_label': source_label,
                                'page_title': page_title,
                                'page_subtitle': page_subtitle,
                            })

                        # Expected columns: material_name, quantity, transaction_type, date
                        rows_imported = 0
                        rows_skipped_duplicate = 0
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if row[0] is None:
                                continue
                            try:
                                material_name = str(row[0]).strip()
                                quantity = int(row[1])
                                transaction_type = _parse_transaction_type(row[2])
                                date_obj = _parse_excel_date(row[3])

                                if not transaction_type or not date_obj:
                                    continue

                                material = Material.objects.filter(name=material_name).first() or \
                                    Material.objects.filter(source_id=material_name).first()

                                if material:
                                    is_duplicate = Transaction.objects.filter(
                                        material=material,
                                        quantity=quantity,
                                        transaction_type=transaction_type,
                                        date=date_obj,
                                    ).exists()

                                    if is_duplicate:
                                        rows_skipped_duplicate += 1
                                    else:
                                        Transaction.objects.create(
                                            material=material,
                                            quantity=quantity,
                                            transaction_type=transaction_type,
                                            date=date_obj
                                        )
                                        rows_imported += 1
                            except Exception:
                                continue

                        result_message = (
                            f'✓ Đã import {rows_imported} bản ghi giao dịch thành công. '
                            f'Bỏ qua {rows_skipped_duplicate} bản ghi trùng hoàn toàn.'
                        )

                except Exception as e:
                    error_message = f'Lỗi xử lý file: {str(e)}'
            else:
                error_message = 'Vui lòng chọn loại dữ liệu và file Excel hợp lệ.'

    return render(request, 'inventory/import_data.html', {
        'form': form,
        'result_message': result_message,
        'error_message': error_message,
        'products': products,
        'materials': materials,
        'min_date': min_date,
        'today': today,
        'sales_source_label': source_label,
        'page_title': page_title,
        'page_subtitle': page_subtitle,
    })


@role_required(ROLE_ADMIN, ROLE_MANAGER, ROLE_STAFF)
def import_data_planning(request):
    return import_data(request, sales_source=SalesData.SOURCE_PLANNING)


@role_required(ROLE_ADMIN, ROLE_MANAGER, ROLE_STAFF)
def multilevel_bom_import(request):
    from .models import MultiLevelBOMEdge
    from .forms import MultiLevelBOMEntryForm

    result_message = None
    error_message = None
    form = MultiLevelBOMEntryForm(request.POST or None)

    if request.method == 'POST':
        # Add new edge via top form
        if 'add_bom_edge' in request.POST:
            if form.is_valid():
                cd = form.cleaned_data
                try:
                    edge, created = MultiLevelBOMEdge.objects.get_or_create(
                        root_product_code=cd['root_product_code'],
                        parent_code=cd['parent_code'],
                        child_code=cd['child_code'],
                        defaults={
                            'parent_type': cd['parent_type'],
                            'child_type': cd['child_type'],
                            'quantity_per_parent': cd['quantity_per_parent'],
                            'level': cd['level'],
                            'remark': cd['remark'],
                        },
                    )
                    if not created:
                        edge.parent_type = cd['parent_type']
                        edge.child_type = cd['child_type']
                        edge.quantity_per_parent = cd['quantity_per_parent']
                        edge.level = cd['level']
                        edge.remark = cd['remark']
                        edge.save()
                        result_message = 'Đã cập nhật cạnh BOM tồn tại.'
                    else:
                        result_message = 'Đã thêm cạnh BOM mới.'
                except Exception as e:
                    error_message = str(e)
            else:
                error_message = 'Dữ liệu nhập không hợp lệ.'

        # Update or delete row via row form
        elif request.POST.get('action_type') == 'update_bom_edge':
            edge_id = request.POST.get('edge_id')
            try:
                edge = MultiLevelBOMEdge.objects.get(id=edge_id)
                edge.parent_code = (request.POST.get('parent_code') or edge.parent_code).upper()
                edge.parent_type = request.POST.get('parent_type') or edge.parent_type
                edge.child_code = (request.POST.get('child_code') or edge.child_code).upper()
                edge.child_type = request.POST.get('child_type') or edge.child_type
                try:
                    edge.quantity_per_parent = float(request.POST.get('quantity_per_parent') or edge.quantity_per_parent)
                except:
                    pass
                try:
                    edge.level = int(request.POST.get('level') or edge.level)
                except:
                    pass
                edge.remark = request.POST.get('remark') or edge.remark
                edge.save()
                result_message = 'Đã lưu thay đổi.'
            except Exception as e:
                error_message = str(e)

        elif request.POST.get('action_type') == 'delete_bom_edge':
            edge_id = request.POST.get('edge_id')
            try:
                MultiLevelBOMEdge.objects.filter(id=edge_id).delete()
                result_message = 'Đã xóa cạnh BOM.'
            except Exception as e:
                error_message = str(e)

    # Build grouped list for template
    edges = MultiLevelBOMEdge.objects.all().order_by('root_product_code', 'level', 'parent_code', 'child_code')
    grouped = {}
    for e in edges:
        grouped.setdefault(e.root_product_code, []).append(e)

    grouped_boms = []
    for root, rows in grouped.items():
        grouped_boms.append({
            'root_product_code': root,
            'row_count': len(rows),
            'rows': rows,
        })

    # ensure form instance for initial display
    if request.method != 'POST':
        form = MultiLevelBOMEntryForm()

    return render(request, 'inventory/import_multilevel_bom.html', {
        'form': form,
        'grouped_boms': grouped_boms,
        'result_message': result_message,
        'error_message': error_message,
    })


# =========================
# MONGODB API ENDPOINTS
# =========================
from .mongodb import (
    insert_material, get_all_materials, get_material_by_id, update_material, delete_material,
    insert_product, get_all_products, get_product_by_id, update_product, delete_product,
    insert_transaction, get_transactions_by_material, update_transaction, delete_transaction
)
from bson.objectid import ObjectId
import json


@session_name_required
def api_mongodb_materials(request):
    """API để quản lý materials trong MongoDB"""
    if request.method == 'GET':
        # Lấy tất cả materials
        materials = get_all_materials()
        materials_list = []
        for mat in materials:
            mat['_id'] = str(mat['_id'])  # Convert ObjectId to string
            materials_list.append(mat)
        return JsonResponse({'status': 'success', 'data': materials_list})
    
    elif request.method == 'POST':
        # Thêm material mới
        try:
            data = json.loads(request.body)
            material_id = insert_material(
                name=data.get('name'),
                on_hand=data.get('on_hand', 0),
                holding_cost=data.get('holding_cost', 0),
                ordering_cost=data.get('ordering_cost', 0),
                price_cost=data.get('price_cost', 0),
                source_id=data.get('source_id')
            )
            return JsonResponse({'status': 'success', 'message': f'Material inserted: {material_id}'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@session_name_required
def api_mongodb_material_detail(request, material_id):
    """API để xem/update/delete material cụ thể"""
    if request.method == 'GET':
        # Lấy material by ID
        material = get_material_by_id(material_id)
        if material:
            material['_id'] = str(material['_id'])
            return JsonResponse({'status': 'success', 'data': material})
        return JsonResponse({'status': 'error', 'message': 'Material not found'}, status=404)
    
    elif request.method == 'PUT':
        # Update material
        try:
            data = json.loads(request.body)
            modified = update_material(material_id, **data)
            return JsonResponse({'status': 'success', 'message': f'Material updated: {modified} documents'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    
    elif request.method == 'DELETE':
        # Delete material
        try:
            deleted = delete_material(material_id)
            return JsonResponse({'status': 'success', 'message': f'Material deleted: {deleted} documents'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@session_name_required
def api_mongodb_products(request):
    """API để quản lý products trong MongoDB"""
    if request.method == 'GET':
        products = get_all_products()
        products_list = []
        for prod in products:
            prod['_id'] = str(prod['_id'])
            products_list.append(prod)
        return JsonResponse({'status': 'success', 'data': products_list})
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            product_id = insert_product(
                name=data.get('name'),
                source_id=data.get('source_id')
            )
            return JsonResponse({'status': 'success', 'message': f'Product inserted: {product_id}'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@session_name_required
def api_mongodb_product_detail(request, product_id):
    """API để xem/update/delete product cụ thể"""
    if request.method == 'GET':
        product = get_product_by_id(product_id)
        if product:
            product['_id'] = str(product['_id'])
            return JsonResponse({'status': 'success', 'data': product})
        return JsonResponse({'status': 'error', 'message': 'Product not found'}, status=404)
    
    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            modified = update_product(product_id, **data)
            return JsonResponse({'status': 'success', 'message': f'Product updated: {modified} documents'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    
    elif request.method == 'DELETE':
        try:
            deleted = delete_product(product_id)
            return JsonResponse({'status': 'success', 'message': f'Product deleted: {deleted} documents'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@session_name_required
def api_mongodb_transactions(request):
    """API để quản lý transactions trong MongoDB"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            transaction_id = insert_transaction(
                material_id=data.get('material_id'),
                transaction_type=data.get('type'),  # 'IN' or 'OUT'
                quantity=data.get('quantity'),
                notes=data.get('notes', '')
            )
            return JsonResponse({'status': 'success', 'message': f'Transaction inserted: {transaction_id}'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@session_name_required
def api_mongodb_material_transactions(request, material_id):
    """API để lấy tất cả transactions của một material"""
    if request.method == 'GET':
        transactions = get_transactions_by_material(material_id)
        transactions_list = []
        for trans in transactions:
            trans['_id'] = str(trans['_id'])
            trans['date'] = trans['date'].isoformat() if hasattr(trans['date'], 'isoformat') else str(trans['date'])
            transactions_list.append(trans)
        return JsonResponse({'status': 'success', 'data': transactions_list})


@session_name_required
def api_mongodb_test(request):
    """Test API - Demo insert, update, delete, query"""
    if request.method == 'POST':
        action = request.POST.get('action')
        
        try:
            if action == 'insert_sample':
                # Insert sample material
                mat_id = insert_material(
                    name='Sample Material',
                    on_hand=100,
                    holding_cost=1.5,
                    ordering_cost=50,
                    price_cost=25.0
                )
                return JsonResponse({'status': 'success', 'message': f'✅ Inserted sample material: {mat_id}'})
            
            elif action == 'get_all':
                # Get all materials
                materials = get_all_materials()
                return JsonResponse({'status': 'success', 'count': len(materials), 'data': [{'_id': str(m['_id']), 'name': m.get('name')} for m in materials]})
            
            elif action == 'update_sample':
                # Update first material
                materials = get_all_materials()
                if materials:
                    first_mat = materials[0]
                    modified = update_material(str(first_mat['_id']), on_hand=200, price_cost=30.0)
                    return JsonResponse({'status': 'success', 'message': f'✅ Updated {modified} materials'})
                return JsonResponse({'status': 'error', 'message': 'No materials found'})
            
            elif action == 'delete_sample':
                # Delete last material
                materials = get_all_materials()
                if materials:
                    last_mat = materials[-1]
                    deleted = delete_material(str(last_mat['_id']))
                    return JsonResponse({'status': 'success', 'message': f'✅ Deleted {deleted} materials'})
                return JsonResponse({'status': 'error', 'message': 'No materials found'})
            
            return JsonResponse({'status': 'error', 'message': 'Invalid action'})
        
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    
    return JsonResponse({'status': 'error', 'message': 'Only POST allowed'})
